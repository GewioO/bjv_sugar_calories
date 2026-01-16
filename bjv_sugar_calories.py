import re
import pandas as pd
import openai
import math
import json
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

with open("keys.json", "r", encoding="utf-8") as f:
    keys = json.load(f)

OPENAI_API_KEY = keys["openai_api_key"]

# Read CSV with presets
def load_known_products(csv_path):
    df = pd.read_csv(csv_path)
    df['назва'] = df['назва'].astype(str)
    return df.set_index('назва')

# Parsing
def parse_full_entry(line):
    pattern = re.compile(
        r'(?P<name>.+?)\s+(?P<kcal>\d+)\s*ккал.*?жири[:\s]*(?P<fat>[\d.]+)[г\s,]+вуглеводи[:\s\-]*(?P<carbs>[\d.]+)[г\s,]+цукри[:\s\-]*(?P<sugars>[\d.]+)[г\s,]+білки[:\s\-]*(?P<protein>[\d.]+)[г\s,]*(?P<weight>[\d.]+)\s*г?', re.IGNORECASE)
    match = pattern.search(line)
    if match:
        return {
            'назва': match.group('name').strip(),
            'ккал': float(match.group('kcal')),
            'жири': float(match.group('fat')),
            'вуглеводи': float(match.group('carbs')),
            'цукри': float(match.group('sugars')),
            'білки': float(match.group('protein')),
            'вага': float(match.group('weight'))
        }
    return None

def get_product_info(name, known_df, input_weight=None):
    name = name.strip().lower()

    for product_name in known_df.index:
        if isinstance(product_name, str) and name in product_name.lower():
            row = known_df.loc[product_name]
            print(f"✅ Знайдено '{name}' як '{product_name}'")

            csv_weight_raw = row.get('вага', '')
            print({csv_weight_raw}, " ", math.isnan(float(float(csv_weight_raw))))
            if not(math.isnan(float(float(csv_weight_raw)))):
                csv_weight = float(csv_weight_raw)
                use_csv_weight = True
            else:
                use_csv_weight = False

            # Визначаємо вагу
            if use_csv_weight:
                final_weight = csv_weight
            elif input_weight is not None:
                final_weight = input_weight
            else:
                final_weight = 100.0

            return {
                'назва': product_name,
                'ккал': float(row['ккал']),
                'жири': float(row['жири']),
                'вуглеводи': float(row['вуглеводи']),
                'цукри': float(row['цукри']),
                'білки': float(row['білки']),
                'вага': final_weight
            }

    print(f"🤖 Звертаємося до GPT для: {name}")
    result = query_gpt_for_nutrition(name)
    if result:
        result['вага'] = input_weight if input_weight else 100.0
    return result

def query_gpt_for_nutrition(product_name):
    prompt = f"""
    Наддай харчову цінність (на 100 г) для продукту '{product_name}' у форматі:
    ккал: <значення>
    жири: <значення>
    вуглеводи: <значення>
    цукри: <значення>
    білки: <значення>
    Без зайвого тексту, лише цифри.
    """

    try:
        response = openai.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
        reply = response.choices[0].message.content.strip()
        print(f"🤖 GPT-відповідь для '{product_name}':\n{reply}")
        pattern = re.compile(
            r'ккал[:\s]*([\d.]+).*?жири[:\s]*([\d.]+).*?вуглеводи[:\s]*([\d.]+).*?цукри[:\s]*([\d.]+).*?білки[:\s]*([\d.]+)',
            re.IGNORECASE | re.DOTALL
        )
        match = pattern.search(reply)
        if match:
            return {
                'назва': product_name,
                'ккал': float(match.group(1)),
                'жири': float(match.group(2)),
                'вуглеводи': float(match.group(3)),
                'цукри': float(match.group(4)),
                'білки': float(match.group(5))
            }
    except Exception as e:
        print(f"❌ GPT-запит не вдався: {e}")
    return None

def parse_simple_entry(line, known_df):
    line = line.strip().lower()
    parts = line.split()
    if not parts:
        return None

    last = parts[-1]
    weight = None
    try:
        weight = eval(last.replace('г', '').replace(',', '.'))
        name = ' '.join(parts[:-1])
    except:
        name = ' '.join(parts)
        weight = None

    product_info = get_product_info(name, known_df, input_weight=weight)
    return product_info

def process_food_log(text_file_path, csv_db_path):
    known_df = load_known_products(csv_db_path)
    final_data = []

    with open(text_file_path, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if not line:
                continue

            data = parse_full_entry(line)
            if not data:
                data = parse_simple_entry(line, known_df)
            if data:
                final_data.append(data)

    return pd.DataFrame(final_data)

def compute_totals(df):
    df = df.copy()
    for col in ['ккал', 'жири', 'вуглеводи', 'цукри', 'білки']:
        df[col + '_факт'] = df[col] * df['вага'] / 100

    totals = df[[c for c in df.columns if '_факт' in c]].sum()
    print("\n📊 Сумарні значення на всі продукти:")
    print(totals.round(2))

    return df

def export_to_excel(df, output_path='nutrition_report.xlsx'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Харчування"
    
    # Стилі
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True, size=11)
    day_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    day_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    headers = ['Index', 'Дата', 'Прийом їжі', 'Страви', 'Білки (г)', 'Жири (г)', 
               'Вуглеводи (г)', 'Цукри (г)', 'Вага (г)', 'Ккал', 'Ккал за прийомом', 'Ккал за день']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    df_sorted = df.sort_values('назва').reset_index(drop=True)
    
    current_row = 2
    row_index = 9
    prev_meal = None
    meal_start_row = 2
    
    for idx, row in df_sorted.iterrows():
        cell = ws.cell(row=current_row, column=1)
        cell.value = row_index
        cell.border = border
        cell.alignment = center_align
        
        cell = ws.cell(row=current_row, column=2)
        cell.value = datetime.now().strftime('%d.%m.%Y')
        cell.border = border
        cell.alignment = center_align
        
        cell = ws.cell(row=current_row, column=3)
        cell.value = "Сніданок"
        cell.border = border
        cell.alignment = center_align
        
        current_meal = "Сніданок" # need to add day method
        
        if prev_meal is not None and prev_meal != current_meal and meal_start_row < current_row:
            formula_cell = ws.cell(row=meal_start_row, column=11)
            formula_cell.value = f'=SUM(J{meal_start_row}:J{current_row - 1})'
            formula_cell.font = Font(bold=True)
            formula_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            meal_start_row = current_row
        
        cell = ws.cell(row=current_row, column=4)
        cell.value = row['назва']
        cell.border = border
        cell.alignment = left_align
        
        cell = ws.cell(row=current_row, column=5)
        cell.value = round(row['білки_факт'], 2)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0.00'
        
        cell = ws.cell(row=current_row, column=6)
        cell.value = round(row['жири_факт'], 2)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0.00'
        
        cell = ws.cell(row=current_row, column=7)
        cell.value = round(row['вуглеводи_факт'], 2)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0.00'
        
        cell = ws.cell(row=current_row, column=8)
        cell.value = round(row['цукри_факт'], 2)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0.00'
        
        cell = ws.cell(row=current_row, column=9)
        cell.value = round(row['вага'], 1)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0.0'
        
        cell = ws.cell(row=current_row, column=10)
        cell.value = round(row['ккал_факт'], 0)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0'
        
        cell = ws.cell(row=current_row, column=11)
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0'
        
        cell = ws.cell(row=current_row, column=12)
        cell.value = 0
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '0'
        
        prev_meal = current_meal
        current_row += 1
        row_index += 1
    
    if meal_start_row < current_row:
        formula_cell = ws.cell(row=meal_start_row, column=11)
        formula_cell.value = f'=SUM(J{meal_start_row}:J{current_row - 1})'
        formula_cell.font = Font(bold=True)
        formula_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 12
    
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"\n✅ Excel-файл збережено: {output_path}")


if __name__ == "__main__":
    df = process_food_log('food_log.txt', 'food_db.csv')

    pd.set_option('display.max_columns', None)      
    pd.set_option('display.width', 1000)            
    pd.set_option('display.max_colwidth', None)   
    
    print("\n🧾 Розпізнаний список продуктів:")
    

    df = compute_totals(df)
    print(df)
    export_to_excel(df, 'nutrition_report.xlsx')

